#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import tkinter as tk
from tkinter import ttk
import pandas as pd
import numpy as np
from PIL import Image
from tkinter import filedialog
import warnings
warnings.filterwarnings("ignore")
import regex
import openpyxl
from openpyxl import load_workbook, formatting, styles
import os
import pycountry
pd.set_option('display.max_columns', None)
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import workbook
from openpyxl.utils import get_column_letter
import customtkinter 
import re
from tkinter import messagebox
from sys import platform
import threading


# In[ ]:


def change_appearance_mode_event(new_appearance_mode: str):
    customtkinter.set_appearance_mode(new_appearance_mode)

def change_scaling_event(new_scaling: str):
    new_scaling_float = int(new_scaling.replace("%", "")) / 100
    customtkinter.set_widget_scaling(new_scaling_float)
    
# Function to choose internal template
def choose_internal_template():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=(("Excel files", ["*.xlsx","*.xlsm"]), ("All files", "*.*")))

def choose_internal_template_1():
    global file_path_1
    file_path_1 = filedialog.askopenfilename(filetypes=(("Excel files", ["*.xlsx","*.xlsm"]), ("All files", "*.*")))

# Function to choose customer template
def choose_customer_template():
    global c_file_path
    c_file_path = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))

def choose_gateway_file():
    global g_file_path
    g_file_path = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))) 
    

def process_customer_template(customer_data, num_rows):
    new_columns = []
    flattened_columns = []

    for i in range(num_rows):
        new_columns.append(customer_data.iloc[i, :].fillna('').astype(str))

    customer_data.columns = pd.MultiIndex.from_frame(pd.concat(new_columns, axis=1))

    main_heading = ""
    temp = {}
    for column in customer_data.columns:
        main_column = column[0]
        sub_columns = []

        if main_column != "":
            main_heading = main_column
            temp = {}

        for j in range(1, num_rows):
            if column[j] != "":
                temp[j] = column[j]
            elif j in temp:
                column = list(column)
                column[j] = temp[j]
                column = tuple(column)
            sub_columns.append(column[j])

        if main_heading == "":
            joined_column = "_".join(sub_columns)
        else:
            joined_column = f'{main_heading}_{("_".join(sub_columns))}'
        flattened_columns.append(joined_column)

    customer_data.columns = flattened_columns
    
    customer_data = customer_data.iloc[num_rows:]
    
    customer_data.reset_index(drop=True, inplace=True)
    return customer_data



def column_treatment(columns):
    
    last_non_unnamed = '' if 'unnamed' in columns[0].lower() else columns[0]  # Initializing last named column
    new_columns = []  # Initializing new list of modified column names
    new_column = ''
    for column in columns:
        if 'unnamed' in column.lower():
            new_column = last_non_unnamed  # Modifying column name to match last named column
        else:
            new_column = column
            last_non_unnamed = new_column  # Updating last named column
        new_columns.append(new_column)

    return new_columns

def template_cleaning(df):
    df.columns = column_treatment(df.columns)
    df.iloc[0] = df.iloc[0].astype(str)
    df.columns = ['_'.join(i) for i in zip(df.columns.get_level_values(0).tolist(), df.iloc[0, :].replace(np.nan, '').tolist())]
    df = df.drop(index=0)
    df.reset_index(drop=True, inplace=True)

    return df


# def transfer_all_data_3():
#     global remaining_columns  # Declare the global variable
#     transferred_columns = [] 
#     remaining_columns = None
#     for i, col in enumerate(o_vars_1):
#         internal_column = internal_template.columns[i]
#         customer_column= col.get()
#         if internal_column and customer_column:
#             internal_template[internal_column] = processed_data[customer_column]
#             transferred_columns.append(customer_column)
#     remaining_columns = processed_data.drop(columns=transferred_columns)        
#     print("Transfer successful")
#     return remaining_columns
#     pass


def template_cleaning_1(df):
    global remaining_columns
    dropped_columns = []
    
    df_without_remaining = df.copy()
    for col in df.columns:
        if col in remaining_columns:
            df_without_remaining.drop(columns=[col], inplace=True)
            dropped_columns.append(col)
    #print(dropped_columns)
    

    # Apply column flattening to df_without_remaining
    df_without_remaining.columns = column_treatment(df_without_remaining.columns)
    df_without_remaining.iloc[0] = df_without_remaining.iloc[0].astype(str)
    df_without_remaining.columns = ['_'.join(i) for i in zip(df_without_remaining.columns.get_level_values(0).tolist(), df_without_remaining.iloc[0, :].replace(np.nan, '').tolist())]
    df_without_remaining = df_without_remaining.drop(index=0)
    df_without_remaining.reset_index(drop=True, inplace=True)
    
    
    # Concatenate df_without_remaining with the original DataFrame containing remaining_columns
    cleaned_df = pd.concat([df_without_remaining, df[dropped_columns]], axis=1)
    #print(cleaned_df.columns)

    return cleaned_df




# In[ ]:


def transfer_data():    
    try:
        num_rows = int(num_rows_entry.get())
    except ValueError:
        tk.messagebox.showerror("Error", "Please enter a valid number of rows.")
        return


    global internal_column_dropdowns
    global processed_data
    global internal_template
    global gcr_og_data
    global gcr_dest_data
    global selected_matches
    global drop_down_states
    #global tab1_internal_columns
    #global ct_columns
   
            
    customer_file_path = c_file_path
    num_rows = int(num_rows_entry.get())
    skip_rows = int(skip_rows_entry.get())
    sheet_name = sheet_entry.get()
    

    if skip_rows > 0:
        skip_rows = skip_rows - 1
        if sheet_name:
            customer_data = pd.read_excel(customer_file_path, skiprows=skip_rows, sheet_name=sheet_name)
        else:
            customer_data = pd.read_excel(customer_file_path, skiprows=skip_rows)
    else:
        if sheet_name:
            customer_data = pd.read_excel(customer_file_path, sheet_name=sheet_name, header=None)
        else:
            customer_data = pd.read_excel(customer_file_path, header=None)

    processed_data = process_customer_template(customer_data, num_rows)
    #ct_columns = list(processed_data.columns)
    

    if file_path:
        internal_template = pd.read_excel(file_path, skiprows=2,sheet_name='Air Internal Template')
        internal_template = template_cleaning(internal_template)
        print("Internal template loaded successfully!")
       
        internal_columns = list(internal_template.columns)
        #tab1_internal_columns = internal_columns.copy()

        internal_column_dropdowns = []
        drop_down_states={}
        #selected_matches = {}

#         def update_selected_matches(column, value):
#             selected_matches[column] = value

        def clear_option_menu(menu_var, default_value):
            menu_var.set(default_value)
            
        remove_b=[]
        o_vars = []
        options = customtkinter.StringVar(frame_2)
        
        for i, column in enumerate(internal_columns):
            label = customtkinter.CTkLabel(frame_2, text=column)
            label.grid(row=i, column=0, sticky="w", padx=10, pady=3)

            dropdown_var = customtkinter.StringVar(frame_2)
            dropdown_var.set("")  # Set initial value to empty string

            dropdown = customtkinter.CTkOptionMenu(frame_2,fg_color="#4D148C",button_color="#4D148C",width=50, variable=dropdown_var, values=list(processed_data.columns))
            dropdown.grid(row=i, column=2, sticky="w", padx=0, pady=10)

            #dropdown.bind('<<ComboboxSelected>>', lambda event, col=column, var=dropdown_var: update_selected_matches(col, var))

            internal_column_dropdowns.append(dropdown)
            
            drop_down_states[column] = dropdown_var
            o_vars.append(dropdown)
        
        for i, col in enumerate(internal_columns):
            b3 = customtkinter.CTkButton(frame_2,fg_color="#4D148C",text='Remove', command=lambda i=i: clear_option_menu(o_vars[i], ""))
            b3.grid(row=i, column=1)
        
    else:
        print("Please select an internal template.")
    pass


# In[ ]:



def transfer_selected_matches():
    global selected_matches
    global drop_down_states
    global s
    s={ v.get():k for k, v in drop_down_states.items()}

    selected_matches = {}
    for i, col in enumerate(internal_template.columns):
        match = internal_column_dropdowns[i].get()
        if match:
            selected_matches[col] = match
            
    #return selected_matches        
    print("Matches Transferred Successfully")
    
    return s
    
   


# In[ ]:


def transfer_data_2():
    if num_rows_entry_.get().strip() == '':
        tk.messagebox.showerror("Error", "Please enter a valid number of rows.")
        return
    
    try:
        num_rows = int(num_rows_entry_.get())
    except ValueError:
        tk.messagebox.showerror("Error", "Please enter a valid number of rows.")
        return

    global internal_column_dropdowns_1
    global processed_data
    global internal_template
    global tab1_internal_columns
    global tab2_internal_columns
    global ct_columns
    global transfer_selected_matches
    global s
    s = {}
    global remaining_columns
    remaining_columns = {}
    customer_file_path = c_file_path
    num_rows = int(num_rows_entry_.get())
    skip_rows = int(skip_rows_entry_.get())
    sheet_name = sheet_entry_.get()
    
    
    if skip_rows > 0:
        skip_rows = skip_rows - 1
        if sheet_name:
            customer_data = pd.read_excel(customer_file_path, skiprows=skip_rows, sheet_name=sheet_name)
        else:
            customer_data = pd.read_excel(customer_file_path, skiprows=skip_rows)
    else:
        if sheet_name:
            customer_data = pd.read_excel(customer_file_path, sheet_name=sheet_name, header=None)
        else:
            customer_data = pd.read_excel(customer_file_path, header=None)
        
    processed_data = process_customer_template(customer_data, num_rows)
    ct_columns = list(processed_data.columns)
    processed_data.iloc[:,:]=''
    
    if file_path_1:
        internal_template_1 = pd.read_excel(file_path_1, skiprows=2,sheet_name='Air Internal Template')
        internal_template_1 = template_cleaning_1(internal_template_1)
        print("Internal template loaded successfully!")
        internal_columns_1 = list(internal_template_1.columns)
        tab1_internal_columns = internal_columns_1.copy()
        #column_index = internal_template_1.shape[1] - 1
        #print(column_index)

    if file_path:
        internal_template = pd.read_excel(file_path, skiprows=None)
        internal_template = template_cleaning_1(internal_template)
        print("Internal template loaded successfully!")
        internal_columns = list(internal_template.columns)
        
        def rem_nan(col):
            if col.endswith("_nan"):
                return col.replace("_nan", "")
            else:
                return col
        
        #internal_columns=[rem_nan(col) for col in internal_template.columns]
        internal_columns = [rem_nan(col) if index > 242 else col for index, col in enumerate(internal_template.columns)]
        internal_template.columns = internal_columns
        #internal_columns = list(internal_template.columns)
        tab2_internal_columns = internal_columns.copy()
        
        internal_column_dropdowns_1 = []
        
        def clear_option_menu(menu_var, default_value):
            menu_var.set(default_value)
            
        remove_b=[]
        o_vars = []
        options = customtkinter.StringVar(frame_2)
        
        for i, col in enumerate(processed_data.columns):
            options = customtkinter.StringVar(frame_2_)
            options.set("")
            label = customtkinter.CTkLabel(frame_2_, text=col)
            label.grid(row=i, column=0, sticky="w", padx=10, pady=3)

            dropdown = customtkinter.CTkOptionMenu(frame_2_, fg_color="#4D148C", button_color="#4D148C", width=50, variable=options, values=list(internal_columns))
            dropdown.set("")
            dropdown.bind('<<ComboboxSelected>>')
            dropdown.grid(row=i, column=2, sticky="w", padx=0, pady=10)
            o_vars.append(options)

            if col in s:
                dropdown.set(s[col])
            else:
                dropdown.set("")

            internal_column_dropdowns_1.append(dropdown)

            
            
            # Pre-fill the dropdown with selected matches from tab 1
#             if column in selected_matches:
#                 dropdown.set(selected_matches[column])
#             else:
#                 dropdown.set("")
                
#             internal_column_dropdowns_1.append(dropdown)
                
        for i, col in enumerate(processed_data.columns):
            b3 = customtkinter.CTkButton(frame_2_,fg_color="#4D148C",text='Remove', command=lambda i=i: clear_option_menu(o_vars[i], ""))
            b3.grid(row=i, column=1)
    
    else:
        print("Please select an internal template.")
    pass
    


# In[ ]:


import json

def load_data_from_file(file_path):
    try:
        with open(file_path, 'r') as file:
            data = json.load(file)
            return data
    except FileNotFoundError:
        return None
    
    
# def load_selected_matches():
#     global selected_matches
#     file_path = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json")])
#     if file_path:
#         selected_matches = load_data_from_file(file_path)
#         if selected_matches:
#             # Update the dropdown selections based on loaded data
#             for col, match in selected_matches.items():
#                 for i, internal_column in enumerate(internal_template.columns):
#                     if internal_column == col:
#                         internal_column_dropdowns_1[i].set(match)
#                         break



def load_selected_matches():
    global selected_matches
    file_path = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json")])
    if file_path:
        with open(file_path, 'r') as json_file:
            selected_matches = json.load(json_file)

        if selected_matches:
            # Reverse the dictionary to set dropdown values in reverse order
            reversed_matches = {v: k for k, v in selected_matches.items()}
            for i, col in enumerate(processed_data.columns):
                if col in reversed_matches:
                    internal_column_dropdowns_1[i].set(reversed_matches[col])

o_vars_1 = []
uploaded_matches = []
def upload_matches():
    global uploaded_matches
    global selected_matches
    global data
    
    def clear_option_menu(menu_var, default_value):
        menu_var.set(default_value)
        
    selected_matches = {}
    file_path = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json")])
    if file_path:
        with open(file_path, 'r') as file:
            data = json.load(file)
            
        for b,(i, k) in enumerate(data.items()):
                label = customtkinter.CTkLabel(frame_2, text=i)
                label.grid(row=b, column=0, sticky="w", padx=10, pady=3)

                dropdown_var = customtkinter.StringVar(frame_2)
                dropdown_var.set(k)  # Set value from uploaded_matches if available
                o_vars_1.append(dropdown_var) # Append the StringVar

                dropdown = customtkinter.CTkOptionMenu(frame_2, width=50, variable=dropdown_var,values=list(processed_data.columns))
                dropdown.grid(row=b, column=2, sticky="w", padx=0, pady=10)
                uploaded_matches.append(k)
                
        for i, col in enumerate(data.items()):
            b3 = customtkinter.CTkButton(frame_2,fg_color="#4D148C",text='Remove', command=lambda i=i: clear_option_menu(o_vars_1[i], ""))
            b3.grid(row=i, column=1)        
                         
               # Update the selected_matches dictionary with values from 'k'
                #selected_matches[k] = i  
        
        messagebox.showinfo("Success", "Matches uploaded successfully.")   
        #return selected_matches   
        
        

def transfer_all_data_3():
    global remaining_columns  # Declare the global variable
    transferred_columns = [] 
    remaining_columns = None
    for i, col in enumerate(o_vars_1):
        internal_column = internal_template.columns[i]
        customer_column= col.get()
        if internal_column and customer_column:
            internal_template[internal_column] = processed_data[customer_column]
            transferred_columns.append(customer_column)
    remaining_columns = processed_data.drop(columns=transferred_columns)        
    print("Transfer successful")
    return remaining_columns
    pass

            
def save_selected_matches():
    file_path = filedialog.asksaveasfilename(defaultextension='.json', filetypes=[("JSON Files", "*.json")])
    if file_path:
        data =  {k: v.get() for k, v in drop_down_states.items()}  # Get values from the tkinter IntVar objects
        
        with open(file_path, 'w') as file:
            json.dump(data, file)
        print("Selected matches saved successfully.")


# Global variable to store column_names data
column_names = []

# Function to load data from a JSON file
def load_data_from_file(file_path):
    global selected_matches, drop_down_states
    if file_path:
        with open(file_path, 'r') as file:
            loaded_data = json.load(file)
        selected_matches = loaded_data.get('selected_matches', {})
        drop_down_states = loaded_data.get('drop_down_states', {})
        # Update dropdowns based on loaded data
        for col, state in loaded_data.items():
            drop_down_states[col] = tk.StringVar(value=state)
            # Update the dropdowns with the loaded data
            if col in column_names:
                drop_downs[col].config(textvariable=drop_down_states[col])
    else:
        messagebox.showerror("Error", "Please choose a valid file.")


# In[ ]:


def country_house_Column(x):
    """This function takes a pandas DataFrame as input and returns a concatenated string of origin port and 
    destination port if both of them are not null, otherwise returns NaN.
    
    Args:
    x (DataFrame): A pandas DataFrame which should have columns named 'origin port' and 'destination port'
    
    Returns:
    str or NaN: A concatenated string of origin port and destination port or NaN
    """
    
    if ((x['origin port'] != 'nan') & (x['destination port'] != 'nan')   ):
        return str(x['origin port']) + " " + str(x['destination port'])
    
    else:
        return 'nan'
    pass


# In[ ]:


def gcr(df,df1,df2):

    df['Origin  _Origin Gateway'] = df['Origin  _Origin Gateway'].astype(str)
    df1.columns = column_treatment(df1.columns)  
    df1.iloc[0] = df1.iloc[0].astype(str)
    df1.columns = ['_'.join(i) for i in zip(df1.columns.get_level_values(0).tolist(), df1.iloc[0,:].replace(np.nan,'').tolist())]
    df1 = df1.drop(index = 0)
    df1.drop_duplicates(subset=['Information_Origin Gateway Code'],keep='first',inplace=True)
    df1['Information_Origin Gateway Code'] = df1['Information_Origin Gateway Code'].astype(str)
    
    #Destination 
    df2.columns = column_treatment(df2.columns)  
    df2.iloc[0] = df2.iloc[0].astype(str)
    df2.columns = ['_'.join(i) for i in zip(df2.columns.get_level_values(0).tolist(), df2.iloc[0,:].replace(np.nan,'').tolist())]
    df2 = df2.drop(index = 0)
    df2.drop_duplicates(subset=['Information_Destination Gateway Code'],keep='first',inplace=True)
    df2['Information_Destination Gateway Code'] = df2['Information_Destination Gateway Code'].astype(str)
 
    df3 = pd.merge(df, df1, left_on = 'Origin  _Origin Gateway', right_on = 'Information_Origin Gateway Code', how = 'left')
    df3 = pd.merge(df3, df2, left_on = 'Destination  _Destination Gateway', right_on = 'Information_Destination Gateway Code', how = 'left')
    df3.to_clipboard()
    df3 = df3.loc[:, :'Dest Remarks_nan']    
    return df3
    


# In[ ]:


def transfer_all_data_2():
    for i, col_dropdown in enumerate(internal_column_dropdowns_1):
        customer_column = processed_data.columns[i]
        internal_column = col_dropdown.get()
        if internal_column and customer_column:
            processed_data[customer_column]=internal_template[internal_column]
            #print(processed_data)
            #print(internal_template)
    print("Transfer successful")
    pass


# In[ ]:


def additional_columns():
    global tab1_internal_columns, tab2_internal_columns, ct_columns
    additional_columns = list(set(tab2_internal_columns) - set(tab1_internal_columns)-set(ct_columns))
    print(additional_columns)
    # Get the index from the processed_data DataFrame
    index = processed_data.index
    
    # Transfer additional columns from internal_template to processed_data
    for column in additional_columns:
        processed_data[column] = internal_template[column]
    
    # Set the index back to the processed_data DataFrame
    processed_data.index = index
    
    print("Additional Columns Transfer successful")


# In[ ]:


# def transfer_all_data():
#     global remaining_columns  # Declare the global variable
#     transferred_columns = [] 
#     remaining_columns = None
#     for i, col_dropdown in enumerate(internal_column_dropdowns):
#         internal_column = internal_template.columns[i]
#         customer_column = col_dropdown.get()
#         if internal_column and customer_column:
#             internal_template[internal_column] = processed_data[customer_column]        
#             transferred_columns.append(customer_column)
         
#     # Create a DataFrame with remaining columns from processed_data
#     remaining_columns = processed_data.drop(columns=transferred_columns)
#     print("Transfer successful")
#     return remaining_columns
#     pass


def transfer_all_data():
    global remaining_columns  
    transferred_columns = [] 
    remaining_columns = None
    
    for i, col_dropdown in enumerate(internal_column_dropdowns):
        internal_column = internal_template.columns[i]
        customer_column = col_dropdown.get()
        
        if not customer_column:
            continue  
        
        if internal_column not in internal_template.columns:
            print(f"Error: The column '{internal_column}' does not exist in 'internal_template'.")
            continue  
        
        if customer_column not in processed_data.columns:
            print(f"Error: The column '{customer_column}' does not exist in 'processed_data'.")
            continue  
        
        try:
            
            internal_template[internal_column] = processed_data[customer_column]
            transferred_columns.append(customer_column)
            
        except Exception as e:
            print(f"Error: An error occurred while transferring data for '{customer_column}': {e}")
    
    
    transferred_columns_set = set(transferred_columns)
    remaining_columns = processed_data[[col for col in processed_data.columns if col not in transferred_columns_set]]
    
    print("Transfer successful")
    return remaining_columns
    pass

def truck_df_treatment():
    states = {
        'AK': 'Alaska',
        'AL': 'Alabama',
        'AR': 'Arkansas',
        'AS': 'American Samoa',
        'AZ': 'Arizona',
        'CA': 'California',
        'CO': 'Colorado',
        'CT': 'Connecticut',
        'DC': 'District of Columbia',
        'DE': 'Delaware',
        'FL': 'Florida',
        'GA': 'Georgia',
        'GU': 'Guam',
        'HI': 'Hawaii',
        'IA': 'Iowa',
        'ID': 'Idaho',
        'IL': 'Illinois',
        'IN': 'Indiana',
        'KS': 'Kansas',
        'KY': 'Kentucky',
        'LA': 'Louisiana',
        'MA': 'Massachusetts',
        'MD': 'Maryland',
        'ME': 'Maine',
        'MI': 'Michigan',
        'MN': 'Minnesota',
        'MO': 'Missouri',
        'MP': 'Northern Mariana Islands',
        'MS': 'Mississippi',
        'MT': 'Montana',
        'NA': 'National',
        'NC': 'North Carolina',
        'ND': 'North Dakota',
        'NE': 'Nebraska',
        'NH': 'New Hampshire',
        'NJ': 'New Jersey',
        'NM': 'New Mexico',
        'NV': 'Nevada',
        'NY': 'New York',
        'OH': 'Ohio',
        'OK': 'Oklahoma',
        'OR': 'Oregon',
        'PA': 'Pennsylvania',
        'PR': 'Puerto Rico',
        'RI': 'Rhode Island',
        'SC': 'South Carolina',
        'SD': 'South Dakota',
        'TN': 'Tennessee',
        'TX': 'Texas',
        'UT': 'Utah',
        'VA': 'Virginia',
        'VI': 'Virgin Islands',
        'VT': 'Vermont',
        'WA': 'Washington',
        'WI': 'Wisconsin',
        'WV': 'West Virginia',
        'WY': 'Wyoming'
}
    global gcr_data
    global internal_template
    global summary_df
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None) 
    df = internal_template
    gateway = pd.read_excel(g_file_path)
    df1 = gateway

    df['Origin  _Origin City'] = df['Origin  _Origin City'].str.upper()
    df['Origin  _Origin State'] = df['Origin  _Origin State'].str.upper()
    df['Destination  _Destination City'] = df['Destination  _Destination City'].str.upper()
    df['Destination  _Destination State'] = df['Destination  _Destination State'].str.upper()
    df['Origin  _Origin State'] = df['Origin  _Origin State'].map(states)
    df['Destination  _Destination State'] = df['Destination  _Destination State'].map(states)

    df['Destination  _Destination Country'].replace('USA', 'US', inplace=True)
    df['Origin  _Origin Country'].replace('USA', 'US', inplace=True)
    pattern=r'([A-Z]){3}'
    #origin gateway correction(removing keyword gateway)       
    df['Origin  _Origin Gateway']=df['Origin  _Origin Gateway'].astype(str).apply(lambda x: re.search(pattern,x)[0] if x and re.search(pattern,x) else np.nan) 
    
        
    #destination gateway keyword removing
    df['Destination  _Destination Gateway'] = df['Destination  _Destination Gateway'].astype(str).apply(lambda x: re.search(pattern, x)[0] if x and re.search(pattern, x) else np.nan)

    # Origin gateway with state
    df1_states = df1[['State', 'Gateway']]
    df1_states['State'] = df1_states['State'].str.upper()
    df1_states.drop_duplicates('State', keep='last', inplace=True)
    df = pd.merge(df, df1_states, left_on='Origin  _Origin State', right_on='State', how='left')
    df['Gateway'] = df['Gateway'].replace('SUB', np.nan)
    df['Origin  _Origin Gateway'].fillna(df['Gateway'], inplace=True)
    #df['Origin  _Origin Gateway'].fillna('', inplace=True)
    del df['State']
    del df['Gateway']

    # Origin gateway with city
    df1_city = df1[[' City/Airport ', 'Gateway']]
    df1_city[' City/Airport '] = df1_city[' City/Airport '].str.upper()
    df1_city.drop_duplicates(' City/Airport ', keep='last', inplace=True)
    df = pd.merge(df, df1_city, left_on='Origin  _Origin City', right_on=' City/Airport ', how='left')
    df['Origin  _Origin Gateway'].fillna(df['Gateway'], inplace=True)
    df['Origin  _Origin Gateway'].fillna('', inplace=True)
    #print(df)
    del df[' City/Airport ']
    del df['Gateway']

    # Destination gateway using states
    df1_states = df1[['State', 'Gateway']]
    df1_states['State'] = df1_states['State'].str.upper()
    df1_states.drop_duplicates('State', keep='last', inplace=True)
    df = pd.merge(df, df1_states, left_on='Destination  _Destination State', right_on='State', how='left')
    df['Gateway'] = df['Gateway'].replace('SUB', np.nan)
    df['Destination  _Destination Gateway'].fillna(df['Gateway'], inplace=True)
    del df['State']
    del df['Gateway']

    # Destination Gateway using city
    df1_city = df1[[' City/Airport ', 'Gateway']]
    df1_city[' City/Airport '] = df1_city[' City/Airport '].str.upper()
    df1_city.drop_duplicates(' City/Airport ', keep='last', inplace=True)
    df = pd.merge(df, df1_city, left_on='Destination  _Destination City', right_on=' City/Airport ', how='left')
    df['Destination  _Destination Gateway'].fillna(df['Gateway'], inplace=True)
    df['Destination  _Destination Gateway'].fillna('', inplace=True)
    del df[' City/Airport ']
    del df['Gateway']

    df1_region = df1[['Region', 'Country']]
    df1_region.drop_duplicates('Country', keep='last', inplace=True)

    # Origin Region
    df = pd.merge(df, df1_region, left_on='Origin  _Origin Country', right_on='Country', how='left')
    df.rename(columns={'Region': 'Origin Region'}, inplace=True)
    del df['Country']

    df = pd.merge(df, df1_region, left_on='Destination  _Destination Country', right_on='Country', how='left')
    df.rename(columns={'Region': 'Destination Region'}, inplace=True)
    del df['Country']

    # Country House column
    df['Origin  _Origin Gateway'] = df['Origin  _Origin Gateway'].astype(str)
    df['Destination  _Destination Gateway'] = df['Destination  _Destination Gateway'].astype(str)
    df['Origin  _Origin Gateway'].replace('no service', 'nan', inplace=True)
    df['Destination  _Destination Gateway'].replace('no service', 'nan', inplace=True)
    df['origin port'] = df['Origin  _Origin Gateway'].copy()
    df['destination port'] = df['Destination  _Destination Gateway'].copy()
    df['Lane Information_Country House'] = df.apply(lambda x: country_house_Column(x), axis=1)
    del df['origin port']
    del df['destination port']

    column_length = df['Origin  _Origin Country'].count()
    df = df.iloc[:column_length, :]
    for index in range(df.shape[0]):
        value = df.loc[index, 'Lane Information_Country House']
        if value.strip() == '':
            df.at[index, 'Country House?'] = 'No'
        else:
            df.at[index, 'Country House?'] = 'Yes'
    print("Gateway Data Transfer Successful!")
    summary_df = df
    df = df.loc[:, :'ALL IN TARGET RATE_COMMENT']
    internal_template = df



# In[ ]:


def summary_tables():
    save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile='Summary Tables')
    df = summary_df
    df1 = df.copy()
    df1=df1.rename(columns={'Lane Information_Country House':'Country House','Lane Information_Lane ID':'No of lanes','Shipment Information_Annual Chargeable KGS':'Sum of Chargeable_Weight Kgs','Shipment Information_# of Shipments':'No of shipments','Origin  _Origin Country':'Origin Country','Destination  _Destination Country':'Destination Country','Origin  _Origin Gateway':'Origin Gateway','Destination  _Destination Gateway':'Destination Gateway','Origin  _Origin City':'Origin City','Destination  _Destination City':'Destination City'})
    #df1.loc[~df1["Country House"].isnull(), 'Country House?'] = 'Yes'
    #df1['Country House?'].fillna('No', inplace = True)
    df1['Hazardous (Y/N)']=df1['Shipment Information_Hazardous (Y/N)'].copy()
    df1['General Cargo'] = df1['Hazardous (Y/N)'].copy()
    df1.loc[df1['General Cargo'] == 'N', 'General Cargo'] = 'No'
    df1['General Cargo'] = df1['General Cargo'].str.upper()
    
    #Region to region Table 
    region_to_region = df1.groupby(['Country House?', 'General Cargo', 'Origin Region',
       'Destination Region']).agg({'No of lanes' : 'count', 'Sum of Chargeable_Weight Kgs' : 'sum', 'No of shipments' : 'sum'}, margins = True)
    try:
        region_to_region['% of Chargeable_Weight'] = (region_to_region['Sum of Chargeable_Weight Kgs'] / region_to_region['Sum of Chargeable_Weight Kgs'].sum()) * 100
    except ZeroDivisionError:
        region_to_region['% of Chargeable_Weight'] = 0
    
    region_to_region['% of Chargeable_Weight'] = region_to_region['% of Chargeable_Weight'].apply(lambda x: round(x, 2))
    try:
        region_to_region['Average weight per shipment (Kgs)'] = (region_to_region['Sum of Chargeable_Weight Kgs'] / (region_to_region['No of shipments'])).astype(int).round()
    except ZeroDivisionError:
        region_to_region['Average weight per shipment (Kgs)'] = 0
        
        
    #Country to country Table
    country_to_country_df = df1.groupby(['Country House?', 'General Cargo', 'Origin Country',
       'Destination Country']).agg({'No of lanes' : 'count', 'Sum of Chargeable_Weight Kgs' : 'sum', 'No of shipments' : 'sum'})
    
    try:
        country_to_country_df['% of Chargeable_Weight'] = (country_to_country_df['Sum of Chargeable_Weight Kgs']/ country_to_country_df['Sum of Chargeable_Weight Kgs'].sum())*100
    except ZeroDivisionError:
        country_to_country_df['% of Chargeable_Weight'] = 0
    country_to_country_df['% of Chargeable_Weight'] = country_to_country_df['% of Chargeable_Weight'].apply(lambda x: round(x, 2))
    try:
        country_to_country_df['Average weight per shipment (Kgs)'] = (country_to_country_df['Sum of Chargeable_Weight Kgs'] / (country_to_country_df['No of shipments'])).astype(int).round()
    except ZeroDivisionError:
        country_to_country_df['Average weight per shipment (Kgs)'] = 0 
        
        
    #Gateway to gateway
    gateway_to_gateway_df = df1.groupby(['Country House?', 'General Cargo', 'Origin Gateway',
       'Destination Gateway']).agg({'No of lanes' : 'count', 'Sum of Chargeable_Weight Kgs' : 'sum', 'No of shipments' : 'sum'})
    try:
        gateway_to_gateway_df['% of Chargeable_Weight'] = (gateway_to_gateway_df['Sum of Chargeable_Weight Kgs']/ gateway_to_gateway_df['Sum of Chargeable_Weight Kgs'].sum())*100
    except ZeroDivisionError:
        gateway_to_gateway_df['% of Chargeable_Weight'] = 0
    gateway_to_gateway_df['% of Chargeable_Weight'] = gateway_to_gateway_df['% of Chargeable_Weight'].apply(lambda x : round(x,2))
    try:
        gateway_to_gateway_df['Average weight per shipment (Kgs)'] = (gateway_to_gateway_df['Sum of Chargeable_Weight Kgs']/(gateway_to_gateway_df['No of shipments'])).astype(int).round()
    except ZeroDivisionError:
        gateway_to_gateway_df['Average weight per shipment (Kgs)'] = 0

    #Origin city to Destination City
    og_city_to_dest_city_df = df1.groupby(['Country House?', 'General Cargo',  'Origin Country',
       'Destination Country','Origin City',
       'Destination City']).agg({'No of lanes' : 'count', 'Sum of Chargeable_Weight Kgs' : 'sum', 'No of shipments' : 'sum'})
    try:
        og_city_to_dest_city_df['% of Chargeable_Weight'] = (og_city_to_dest_city_df['Sum of Chargeable_Weight Kgs']/ og_city_to_dest_city_df['Sum of Chargeable_Weight Kgs'].sum())*100
    except ZeroDivisionError:
        og_city_to_dest_city_df['% of Chargeable_Weight'] = 0
    og_city_to_dest_city_df['% of Chargeable_Weight'] = og_city_to_dest_city_df['% of Chargeable_Weight'].apply(lambda x: round(x, 2))
    try:
        og_city_to_dest_city_df['Average weight per shipment (Kgs)'] = (og_city_to_dest_city_df['Sum of Chargeable_Weight Kgs']/(og_city_to_dest_city_df['No of shipments'])).astype(int).round()
    except ZeroDivisionError:
        og_city_to_dest_city_df['Average weight per shipment (Kgs)'] = 0 

    #Origin
    origin_df = df1.groupby([  'Origin Country',
      'Origin Gateway','Origin City',
       ]).agg({'No of lanes' : 'count', 'Sum of Chargeable_Weight Kgs' : 'sum', 'No of shipments' : 'sum'})
    
    try:
        origin_df['% of Chargeable_Weight'] = (origin_df['Sum of Chargeable_Weight Kgs']/ origin_df['Sum of Chargeable_Weight Kgs'].sum())*100
    except ZeroDivisionError:
        origin_df['% of Chargeable_Weight'] = 0 
    origin_df['% of Chargeable_Weight'] = origin_df['% of Chargeable_Weight'].apply(lambda x: round(x, 2))
    try:
        origin_df['Average weight per shipment (Kgs)'] = (origin_df['Sum of Chargeable_Weight Kgs']/(origin_df['No of shipments'] )).astype(int).round()
    except ZeroDivisionError:
        origin_df['Average weight per shipment (Kgs)'] = 0 
        
    #Destination
    destination_df = df1.groupby([  'Destination Country',
      'Destination Gateway','Destination City',
       ]).agg({'No of lanes' : 'count', 'Sum of Chargeable_Weight Kgs' : 'sum', 'No of shipments' : 'sum'})
    try:
        destination_df['% of Chargeable_Weight'] = (destination_df['Sum of Chargeable_Weight Kgs']/ destination_df['Sum of Chargeable_Weight Kgs'].sum())*100
    except ZeroDivisionError:
        destination_df['% of Chargeable_Weight'] = 0
    destination_df['% of Chargeable_Weight'] = destination_df['% of Chargeable_Weight'].apply(lambda x : round(x,2))
    try:
        destination_df['Average weight per shipment (Kgs)'] = (destination_df['Sum of Chargeable_Weight Kgs']/(destination_df['No of shipments'])).astype(int).round()
    except ZeroDivisionError:
        destination_df['Average weight per shipment (Kgs)'] = 0

    with pd.ExcelWriter(save_path) as writer:
        
    # Write each dataframe to a separate sheet
        region_to_region.to_excel(writer, sheet_name='Region to Regrion')
        country_to_country_df.to_excel(writer, sheet_name='Country to Country')
        gateway_to_gateway_df.to_excel(writer, sheet_name='Gateway to Gateway')
        og_city_to_dest_city_df.to_excel(writer, sheet_name='Origin city to Destion city')
        origin_df.to_excel(writer, sheet_name='Origin')
        destination_df.to_excel(writer, sheet_name='Destination')  
    print("Summary Table saved successfully.")


# In[ ]:


def download_excel():
    file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
    new_columns = [tuple(col.split('_')) for col in  internal_template.columns]
    internal_template.columns = pd.MultiIndex.from_tuples(new_columns)
    internal_template.dropna(how='all',inplace=True)   
    internal_template.reset_index(drop=True, inplace=True)
    l=len(internal_template)
    index=pd.Index(range(l),name='index')
    internal_template.index=index
    internal_template.to_excel(file_path)
    start_col = len(internal_template.columns) + 2 
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    for c_idx, (col_name, values) in enumerate(remaining_columns.items()):
        sheet.cell(row=1, column=start_col + c_idx, value=col_name)  # Write column name
    
         # Iterate over values and write them in the same column
        for r_idx, value in enumerate(values):
            sheet.cell(row=r_idx + 4, column=start_col + c_idx, value=value)



    # Set the row height for the first row to 73.5
    sheet.row_dimensions[1].height = 73.5

    # Set the row height for the second row to 48
    sheet.row_dimensions[2].height = 48
    
    # Set wrap text for the first row
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    # Apply formatting to row 1
    font_row1 = Font(name='Calibri', size=14, bold=True)
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = font_row1
            
    # Apply formatting to row 2
    font_row2 = Font(name='Calibri', size=12, bold=True)
    for row in sheet.iter_rows(min_row=2, max_row=2):
        for cell in row:
            cell.font = font_row2

    # Colors for cells
    color_cells(sheet, 'B1:G1', 'D9D9D9')
    color_cells(sheet, 'B2:G2', 'D9D9D9')
    color_cells(sheet, 'H1:N1' ,'95B3D7')
    color_cells(sheet, 'H2:N2' ,'95B3D7')
    color_cells(sheet, 'O1:U1' ,'B1A0C7')
    color_cells(sheet, 'O2:U2' ,'B1A0C7')
    color_cells(sheet, 'V1:X1' ,'BFBFBF')
    color_cells(sheet, 'V2:X2' ,'BFBFBF')
    color_cells(sheet, 'Y1:AJ1' ,'BFBFBF')
    color_cells(sheet, 'Y2:AJ2' ,'BFBFBF')
    color_cells(sheet, 'AK1:DD1' ,'95B3D7')
    color_cells(sheet, 'AK2:DD2' ,'95B3D7')
    color_cells(sheet, 'DE1:FF1' ,'BFBFBF')
    color_cells(sheet, 'DE2:FF2' ,'BFBFBF')
    color_cells(sheet, 'FG1:HY1' ,'B1A0C7')
    color_cells(sheet, 'FG2:HY2' ,'B1A0C7')
    color_cells(sheet, 'HZ1:II1' ,'E6B8B7')
    color_cells(sheet, 'HZ2:II2' ,'E6B8B7')
    
    # Adjust column widths based on the content
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            if cell.coordinate in sheet.merged_cells:
                continue
            value = cell.value
            if value:
                cell_length = len(str(value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = (max_length + 2) * 1.2
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    
#     columns_to_check = ['C', 'D', 'E', 'F']

#     # Function to check if a column is completely blank
#     def is_column_blank(column_letter):
#         for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=ord(column_letter) - ord('A') + 1, max_col=ord(column_letter) - ord('A') + 1):
#             for cell in row:
#                 if cell.value is not None:
#                     return False
#         return True

#     # Hide the specified columns if they are blank
#     for column in columns_to_check:
#         if is_column_blank(column):
#             sheet.column_dimensions[column].hidden = True

    
        
    # Save the updated workbook
    workbook.save(file_path)
    print("Excel Saved Successfully")
    pass



# In[ ]:


def color_cells(sheet, cell_range, fill_color):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = fill


# In[ ]:


def download_excel_2():
    global global_destination_file
    file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
    new_columns = [tuple(col.split('_')) for col in  processed_data.columns]
    processed_data.columns = pd.MultiIndex.from_tuples(new_columns)
    processed_data.dropna(how='all',inplace=True)   
    processed_data.reset_index(drop=True, inplace=True)
    l=len(processed_data)
    index=pd.Index(range(l),name='index')
    processed_data.index=index
    processed_data.to_excel(file_path)
    global_destination_file = file_path
    pass


# In[ ]:


# def customer_template_format():
#     global global_destination_file
#     global c_file_path
    
#     source_start_row = int(num_rows_entry_1.get())
#     destination_start_row = int(skip_rows_entry_1.get())
#     sheet_name_1 = sheet_entry.get()
    



#     # Load the source and destination workbooks
#     dest_wb = openpyxl.load_workbook(c_file_path)
#     source_wb = openpyxl.load_workbook(global_destination_file)

#     # Get the first sheet from each workbook
#     source_worksheet = source_wb.active
#     if not sheet_name_1:
#         destination_worksheet = dest_wb.active  # Use the default active sheet
#     else:
#         destination_worksheet = dest_wb[sheet_name_1]

#    # Delete the first column (column A) in the source worksheet
#     source_worksheet.delete_cols(1)

#  # Clear the destination data, formatting, and styles from the specified row onwards
#     for row_index in range(destination_start_row, destination_worksheet.max_row + 1):
#         for col_index in range(1, destination_worksheet.max_column + 1):
#             destination_worksheet.cell(row=row_index, column=col_index, value=np.nan)
#             destination_worksheet.cell(row=row_index, column=col_index).font = openpyxl.styles.Font()
#             destination_worksheet.cell(row=row_index, column=col_index).border = openpyxl.styles.Border()
#             destination_worksheet.cell(row=row_index, column=col_index).fill = openpyxl.styles.PatternFill()

#     # Copy data from the source to the destination worksheet
#     for row_index, source_row in enumerate(source_worksheet.iter_rows(min_row=source_start_row), start=destination_start_row):
#         for col_index, cell in enumerate(source_row, start=1):
#             destination_worksheet.cell(row=row_index, column=col_index, value=cell.value)

#     # Save the destination workbook
#     dest_wb.save(c_file_path)
#     print("Trasnfer Successfull")


# In[ ]:


def get_source_column_names(source_worksheet, source_br_column_index):
    source_column_names = []
    add_columns = False

    for col_index, cell in enumerate(source_worksheet[1], start=1):
        if add_columns:
            source_column_names.append(cell.value)
        if col_index == source_br_column_index+1:
            add_columns = True

    print(source_column_names)        
    return source_column_names

def customer_template_format():
    global global_destination_file
    global c_file_path
    

    source_start_row = int(num_rows_entry_1.get())
    destination_start_row = int(skip_rows_entry_1.get())
    sheet_name_1 = sheet_entry.get()
    skip_rows = int(skip_rows_entry_.get())
    print(skip_rows)

    # Load the source and destination workbooks
    dest_wb = openpyxl.load_workbook(c_file_path)
    source_wb = openpyxl.load_workbook(global_destination_file)

    # Get the first sheet from each workbook
    source_worksheet = source_wb.active
    if not sheet_name_1:
        destination_worksheet = dest_wb.active  # Use the default active sheet
    else:
        destination_worksheet = dest_wb[sheet_name_1]

    # Get the index of the last column
    last_column_index = destination_worksheet.max_column - 1
    print(last_column_index)
    # Delete the first column (column A) in the source worksheet
    source_worksheet.delete_cols(1)

    # Clear the destination data, formatting, and styles from the specified row onwards
    for row_index in range(destination_start_row, destination_worksheet.max_row + 1):
        for col_index in range(1, destination_worksheet.max_column + 1):
            destination_worksheet.cell(row=row_index, column=col_index, value=np.nan)
            destination_worksheet.cell(row=row_index, column=col_index).font = openpyxl.styles.Font()
            destination_worksheet.cell(row=row_index, column=col_index).border = openpyxl.styles.Border()
            destination_worksheet.cell(row=row_index, column=col_index).fill = openpyxl.styles.PatternFill()

    for row_index, source_row in enumerate(source_worksheet.iter_rows(min_row=source_start_row), start=destination_start_row):
        for col_index, cell in enumerate(source_row, start=1):
            destination_worksheet.cell(row=row_index, column=col_index, value=cell.value)

    #source_br_column_index = last_column_index("BP")

    if last_column_index is not None:
      
        destination_start_column = last_column_index + 1

       
        source_column_names = get_source_column_names(source_worksheet, last_column_index)
        for col_index, column_name in enumerate(source_column_names, start=destination_start_column+1):
            destination_worksheet.cell(row=skip_rows+ 1, column=col_index, value=column_name)
    else:
        print("Column not found in the source worksheet.")

    # Save the destination workbook
    dest_wb.save(c_file_path)
    print("Transfer Successful")


# In[ ]:


# Function to open the main UI when the user logs in
def open_main_ui():
    # Close the login window
    root.destroy()


# In[ ]:


root = customtkinter.CTk()
width=root.winfo_screenwidth()
height=root.winfo_screenheight()
# window.geometry("%dx%d"%(w,h))


root.title("Login Page")

# x = (root.winfo_screenwidth() // 2) - (width //2 )
# y = (root.winfo_screenwidth() // 3) - (height //2 )

# root.geometry('{} x {} + {}'.format(width,height,x,y))
root.geometry("%dx%d"%(width,height))
root.resizable(True,True)


bg_image = customtkinter.CTkImage(Image.open('background_image.png'), size = (width,height))
bg_image_label = customtkinter.CTkLabel(root, image = bg_image, text = "" )
bg_image_label.grid(row = 0, column = 0, padx = (0,120), pady = (0,70))


login_btn = customtkinter.CTkButton(root, text = "START", width = 200, height = 60,command = open_main_ui)
login_btn.grid(row = 0, column = 0, padx= (0,600), pady = (220,100))


root.mainloop()


# In[ ]:


import tkinter as tk
from PIL import Image, ImageTk



# Create the main window
global transfer_data
customtkinter.set_appearance_mode("light")
customtkinter.set_default_color_theme("dark-blue")
window = customtkinter.CTk()
my_image = customtkinter.CTkImage(light_image=Image.open("Pict.png"),
                                  dark_image=Image.open("Pict.png"),
                                  size=(60,30))
window.wm_iconbitmap('Picture3.ico')
#window.geometry("1536x864")
window.title("Bid Processing App.py")
#window.attributes('-fullscreen',True)
w=window.winfo_screenwidth()
h=window.winfo_screenheight()
window.geometry("%dx%d"%(w,h))

window.state("zoomed")
window.grid_rowconfigure((0,1,2), weight=1)
window.grid_columnconfigure((0,1,2), weight=1)

# bg_image1 = customtkinter.CTkImage(Image.open('Pict - Copy.png'), size = (width,height))
# bg_image_label_1 = customtkinter.CTkLabel(window, image = bg_image1, text = "" )
# bg_image_label_1.grid(row = 0, column = 0, padx = (0,120), pady = (0,70))


tabview = customtkinter.CTkTabview(master=window,border_width=2,border_color="#FF6600")
#tabview.grid_rowconfigure((0,1,2), weight=1)
tabview.grid(row=0,sticky='nsew',padx=20,pady=20,rowspan=3,columnspan=3)

# Create a Canvas for Pre Analysis
# canvas_tab1 = tk.Canvas(tabview.get_tab("Pre Analysis"))



tabview.add("Pre Analysis")  # add tab at the end
tabview.add("Post Analysis")  # add tab at the end
tabview.set("Pre Analysis")  # set currently visible tab



button = customtkinter.CTkButton(master=tabview.tab("Post Analysis"))
button.grid(row=1,sticky='nsew',padx=20, pady=20)


frame_1 = customtkinter.CTkFrame(master=tabview.tab("Pre Analysis"))
frame_1.grid(row=0,sticky='nsew',padx=20,pady=25)

frame_2 = customtkinter.CTkScrollableFrame(master=tabview.tab("Pre Analysis"), orientation="vertical",width=950, height=800)
frame_2.grid(row=0, column=1, padx=30, pady=20,sticky='nsew',rowspan=9)
frame_3=customtkinter.CTkFrame(master=tabview.tab("Pre Analysis"))
frame_3.grid(row=1,column=0,sticky='nsew',padx=20,pady=25)
frame_4=customtkinter.CTkFrame(master=tabview.tab("Pre Analysis"))
frame_4.grid(row=2,sticky='nsew',padx=20,pady=25)
frame_5=customtkinter.CTkFrame(master=tabview.tab("Pre Analysis"))
frame_5.grid(row=3,sticky='nsew',padx=20,pady=25)
frame_6=customtkinter.CTkFrame(master=tabview.tab("Pre Analysis"))
frame_6.grid(row=4,sticky='nsew',padx=20,pady=25)


frame_1_ = customtkinter.CTkFrame(master=tabview.tab("Post Analysis"))
frame_1_.grid(row=1,sticky='nsew',padx=20,pady=20)
frame_2_ = customtkinter.CTkScrollableFrame(master=tabview.tab("Post Analysis"), orientation="vertical",width=950, height=800)
frame_2_.grid(row=0, column=1, padx=30, pady=20,sticky='nsew',rowspan=9)
frame_3_=customtkinter.CTkFrame(master=tabview.tab("Post Analysis"))
frame_3_.grid(row=2,sticky='nsew',padx=20,pady=20)
frame_4_=customtkinter.CTkFrame(master=tabview.tab("Post Analysis"))
frame_4_.grid(row=3,sticky='nsew',padx=20,pady=20)
frame_5_=customtkinter.CTkFrame(master=tabview.tab("Post Analysis"))
frame_5_.grid(row=4,sticky='nsew',padx=20,pady=20)


heading_i=customtkinter.CTkLabel(tabview,image=my_image,text="",text_color="#999999",font=("Arial",16,"bold"))
heading_i.grid(row=0, column=0, sticky="w", padx=10, pady=3)

# ===============================Button Image=================================================
button_image = customtkinter.CTkImage(Image.open("excelimage1.png"), size=(26, 26))
transfer_image = customtkinter.CTkImage(Image.open("transfer_image.png"), size=(26, 26))
download_image = customtkinter.CTkImage(Image.open("downloadimage.png"), size=(26, 26))


# ===========================================================================================


# Choose Internal Template button
internal_template_button_ = customtkinter.CTkButton(frame_1_,fg_color="#4D148C", text="Downloaded Template", command=choose_internal_template,
                                     width=28, anchor="center", image = button_image)
internal_template_button_.grid(row=1, column=1, sticky="w", padx=10, pady=3)

internal_template_button_og = customtkinter.CTkButton(frame_1_,fg_color="#4D148C", text="Original Template", command=choose_internal_template_1,
                                     width=28, anchor="center", image = button_image)
internal_template_button_og.grid(row=1, column=0, sticky="w", padx=10, pady=3)

# Choose Customer Template button
customer_template_button_ = customtkinter.CTkButton(frame_1_,fg_color="#4D148C", text="Choose Customer Template", command=choose_customer_template,
                                     width=28, anchor="center", image = button_image)
customer_template_button_.grid(row=1, column=2, sticky="w", padx=10, pady=3)


# Label and Entry for number of rows to skip
label_s_entry_ = customtkinter.CTkLabel(frame_3_, text="Number of Rows to Skip:")
label_s_entry_.grid(row=2, column=0, sticky="w",padx=10, pady=3)

skip_rows_entry_ = customtkinter.CTkEntry(frame_3_, width=34)
skip_rows_entry_.grid(row=2, column=1, sticky="w", padx=10, pady=3)

# Label and Entry for number of headers
label_n_entry_ = customtkinter.CTkLabel(frame_3_, text="Number of Rows to Combine:")
label_n_entry_.grid(row=3, column=0, sticky="w",padx=10, pady=3)

num_rows_entry_ = customtkinter.CTkEntry(frame_3_, width=34)
num_rows_entry_.grid(row=3, column=1, sticky="w", padx=10, pady=3)

# Label and Entry for sheet name
label_sheet_entry_ = customtkinter.CTkLabel(frame_3_, text="Sheet Name:")
label_sheet_entry_.grid(row=4, column=0, sticky="w",padx=10, pady=3)

sheet_entry_ = customtkinter.CTkEntry(frame_3_, width=100)
sheet_entry_.grid(row=4, column=1, sticky="w", padx=10, pady=20)

# transfer_matches_button = customtkinter.CTkButton(frame_3_, fg_color="#4D148C", text="Load Matches", command=load_selected_matches, width=28, anchor="center")
# transfer_matches_button.grid(row=5, column=1, sticky="w", padx=10, pady=3)


# Show Columns button
# transfer_button_ = customtkinter.CTkButton(frame_4_,fg_color="#4D148C", text="Show Columns", command=transfer_data_2, width=28, anchor="center")
# transfer_button_.grid(row=7, column=0, sticky="w", padx=10, pady=3)

transfer_button_ = customtkinter.CTkButton(frame_4_, fg_color="#4D148C", text="Show Columns", command = threading.Thread(target=transfer_data_2).start, width=28, anchor="center")
transfer_button_.grid(row=7, column=0, sticky="w", padx=10, pady=3)


# transfer_button = customtkinter.CTkButton(frame_3,fg_color="#4D148C", text="RFQ Attributes", command=transfer_data, width=28, anchor="center")
# transfer_button.grid(row=0, column=0, sticky="nsew", padx=10, pady=3)

 # Transfer button for all columns
transfer_all_button = customtkinter.CTkButton(frame_4_,fg_color="#4D148C", text="Transfer All Data", command=transfer_all_data_2, width=28, anchor="center",image =  transfer_image)
transfer_all_button.grid(row=7, column=1, sticky="w", padx=10, pady=3)



transfer_all_button = customtkinter.CTkButton(frame_4_,fg_color="#4D148C", text=" Transfer Additonal Columns", command=additional_columns, width=28, anchor="center",image =  transfer_image)
transfer_all_button.grid(row=7, column=2, sticky="w", padx=10, pady=3)
# # Transfer Gateway Data button
# transfer_all_button = customtkinter.CTkButton(frame_4_, text="Transfer Gateway Data", command=truck_df_treatment, width=28, anchor="center")
# transfer_all_button.grid(row=7, column=2, sticky="w", padx=10, pady=3)

download_button = customtkinter.CTkButton(frame_4_,fg_color="#4D148C", text="Download Customer Template", command=download_excel_2, width=28, anchor="center", image = download_image)
download_button.grid(row=8, column=0, sticky="w", padx=10, pady=3)

# Label and Entry for number of headers
label_n_entry_ = customtkinter.CTkLabel(frame_5_, text="Enter the row number to start copying from:")
label_n_entry_.grid(row=3, column=0, sticky="w")

num_rows_entry_1 = customtkinter.CTkEntry(frame_5_, width=34)
num_rows_entry_1.grid(row=3, column=1, sticky="w", padx=10, pady=3)

# Label and Entry for number of rows to skip in final customer template
label_n_entry_ = customtkinter.CTkLabel(frame_5_, text="Enter the row number to start pasting to:")
label_n_entry_.grid(row=4, column=0, sticky="w")

skip_rows_entry_1 = customtkinter.CTkEntry(frame_5_, width=34)
skip_rows_entry_1.grid(row=4, column=1, sticky="w", padx=10, pady=3)

download_button = customtkinter.CTkButton(frame_5_,fg_color="#4D148C",text="Update Customer Template", command=customer_template_format, width=28, anchor="center")
download_button.grid(row=5, column=0, sticky="w", padx=10, pady=3)



# heading_l=tk.Label(window,text="Bid Processing App",font=("Arial",16,"bold"))
# heading_l.grid(row=0, column=0, sticky="w", padx=10, pady=3)

#TAB1

# Choose Internal Template button
internal_template_button = customtkinter.CTkButton(frame_1,fg_color="#4D148C", text="Internal Template",border_color="#FF6600",border_width=2, command=choose_internal_template,
                                     width=28, anchor="center", image = button_image)
internal_template_button.grid(row=1, column=0, sticky="w", padx=10, pady=3)

# Choose Customer Template button
customer_template_button = customtkinter.CTkButton(frame_1,fg_color="#4D148C", text="Customer Template",border_color="#FF6600",border_width=2, command=choose_customer_template,
                                     width=28, anchor="center", image = button_image)
customer_template_button.grid(row=1, column=1, sticky="w", padx=10, pady=3)


#gateway Button

customer_template_button = customtkinter.CTkButton(frame_1,fg_color="#4D148C", text="Gateway", border_color="#FF6600",border_width=2,
                                                   command=choose_gateway_file,
                                     width=28, anchor="center", image = button_image)
customer_template_button.grid(row=1, column=2, sticky="w", padx=10, pady=3)

# Label and Entry for number of rows to skip
label_s_entry = customtkinter.CTkLabel(frame_1, text="Rows to Skip:")
label_s_entry.grid(row=2, column=0, sticky="w",padx=10, pady=1)

skip_rows_entry = customtkinter.CTkEntry(frame_1,border_color="#FF6600",border_width=2, width=34)
skip_rows_entry.grid(row=2, column=1, sticky="w",padx=10, pady=1 )

# Label and Entry for number of headers
label_n_entry = customtkinter.CTkLabel(frame_1, text="Rows to Combine:")
label_n_entry.grid(row=2, column=2, sticky="w",padx=10, pady=3)

num_rows_entry = customtkinter.CTkEntry(frame_1,border_color="#FF6600",border_width=2, width=34)
num_rows_entry.grid(row=2, column=3, sticky="w", padx=10, pady=3)

# Label and Entry for sheet name
label_sheet_entry = customtkinter.CTkLabel(frame_1, text="Sheet Name:")
label_sheet_entry.grid(row=3, column=0, sticky="w",padx=10, pady=3)

sheet_entry = customtkinter.CTkEntry(frame_1,border_color="#FF6600",border_width=2, width=100)
sheet_entry.grid(row=3, column=1, sticky="w", padx=10, pady=3)

# Show Columns button

transfer_button = customtkinter.CTkButton(frame_3,fg_color="#4D148C", text="RFQ Attributes", command = threading.Thread(target=transfer_data).start, width=28, anchor="center")
transfer_button.grid(row=0, column=0, sticky="nsew", padx=10, pady=3)

 # Transfer button for all columns
transfer_all_button = customtkinter.CTkButton(frame_3,fg_color="#4D148C", text="Convert(CT-IT)", command=threading.Thread(target=transfer_all_data).start , width=28, anchor="center", image = transfer_image)
transfer_all_button.grid(row=0, column=1, sticky="nsew", padx=10, pady=3)

# Transfer Gateway Data button
transfer_all_button = customtkinter.CTkButton(frame_3,fg_color="#4D148C", text="Update Gateway", command=truck_df_treatment, width=28, anchor="center", image = transfer_image)
transfer_all_button.grid(row=0, column=2, sticky="nsew", padx=10, pady=3)

# Download Summary Excel button
download_button = customtkinter.CTkButton(frame_4,fg_color="#4D148C", text="Summary Table", command=summary_tables, width=28, anchor="center", image = download_image)
download_button.grid(row=2, column=0, sticky="ew", padx=10, pady=3)

# Download Excel button
download_button = customtkinter.CTkButton(frame_4,fg_color="#4D148C", text="Internal Template", command=download_excel, width=28, anchor="center", image = download_image)
download_button.grid(row=2, column=1, sticky="w", padx=10, pady=3)


# # Create a new button in "Frame2" of "Pre Analysis" to transfer selected matches
# transfer_matches_button = customtkinter.CTkButton(frame_3_,fg_color="#4D148C", text="Transfer Matches", command=transfer_selected_matches, width=28, anchor="center")
# transfer_matches_button.grid(row=5, column=1, sticky="w", padx=10, pady=3)
transfer_matches_button = customtkinter.CTkButton(frame_3_, fg_color="#4D148C", text="Load Matches", command=load_selected_matches, width=28, anchor="center")
transfer_matches_button.grid(row=5, column=1, sticky="w", padx=10, pady=3)



# Upload Matches button
upload_matches_button = customtkinter.CTkButton(frame_5,fg_color="#4D148C", text="Upload", command=upload_matches, width=28, anchor="center")
upload_matches_button.grid(row=0, column=0, sticky="ew", padx=10, pady=3)

d_button = customtkinter.CTkButton(frame_5,fg_color="#4D148C", text="RFQ Mapping", command=transfer_all_data_3, width=28, anchor="center")
d_button.grid(row=0, column=1, sticky="w", padx=10, pady=3)

# Download Matches button
download_matches_button = customtkinter.CTkButton(frame_5,fg_color="#4D148C", text="Store Mapping", command=save_selected_matches, width=28, anchor="center")
download_matches_button.grid(row=0, column=2, sticky="w", padx=10, pady=3)

appearance_mode_optionemenu = customtkinter.CTkOptionMenu(frame_6,fg_color="#4D148C" ,values=["Light" ,"Dark","System"],anchor="center" ,command=change_appearance_mode_event,width=28)
appearance_mode_optionemenu.grid(row=0, column=2,sticky="nsew", padx=10, pady=3)

scaling_optionemenu = customtkinter.CTkOptionMenu(frame_6,fg_color="#4D148C", values=["80%", "85%","90%","100%", "110%", "120%"], anchor="center" ,command=change_scaling_event,width=28)
scaling_optionemenu.grid(row=0, column=3,sticky="nsew", padx=10, pady=3)
scaling_optionemenu.set("80%")





# Run the main window loop
window.mainloop()     

# # Create the login window
# login_window = tk.Tk()
# login_window.title("Login")
# login_window.geometry("400x200")

# # Add a label for the username
# username_label = tk.Label(login_window, text="Username:")
# username_label.pack()

# # Add an entry widget for the username
# username_entry = tk.Entry(login_window)
# username_entry.pack()

# # Add a label for the password
# password_label = tk.Label(login_window, text="Password:")
# password_label.pack()

# # Add an entry widget for the password (you can replace this with a password entry widget for security)
# password_entry = tk.Entry(login_window, show="*")
# password_entry.pack()

# # Add a button to log in and open the main UI
# login_button = tk.Button(login_window, text="Login", command=open_main_ui)
# login_button.pack()

# # Add any other widgets you need for the login screen

# # Start the main loop for the login window
# login_window.mainloop()




# In[ ]:





# In[ ]:





# In[ ]:




